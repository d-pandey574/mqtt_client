import sys
import time
import json
import serial
import paho.mqtt.client as mqtt
from serial.tools.list_ports import comports
from openpyxl import Workbook, load_workbook

# MQTT Configuration
broker_address = "mqtt.eclipseprojects.io"
port = 1883
client_id = "mqttx_934f8f39"
publish_topic = "sensor/data"
subscribe_topic = "sensor/cammond"

# Excel file path
excel_file = "sensor_data7.xlsx"

# MQTT Callbacks
def on_connect(client, userdata, flags, rc):
    print(f"[MQTT] Connected with result code {rc}")
    client.subscribe(subscribe_topic)
    print(f"[MQTT] Subscribed to {subscribe_topic}")

def on_message(client, userdata, msg):
    print(f"[MQTT] Received message: {msg.payload.decode()} on topic: {msg.topic}")
    # Handle storing the data
    store_data_in_excel(json.loads(msg.payload.decode()))

# Initialize MQTT
client = mqtt.Client(client_id)
client.on_connect = on_connect
client.on_message = on_message
client.connect(broker_address, port, 60)
client.loop_start()

# List available COM ports
def ask_for_port():
    ports = []
    print("\nAvailable COM Ports:")
    for i, (port, desc, _) in enumerate(sorted(comports()), 1):
        print(f"{i}: {port} ({desc})")
        ports.append(port)
    while True:
        choice = input("Enter port number or full name: ")
        try:
            index = int(choice) - 1
            if 0 <= index < len(ports):
                return ports[index]
            else:
                print("Invalid index!")
        except ValueError:
            if choice in ports:
                return choice
            print("Invalid input!")

# Function to convert HEX string to JSON
def hex_to_json(hex_str):
    try:
        ascii_str = bytes.fromhex(hex_str).decode('utf-8')
        return json.loads(ascii_str)
    except Exception as e:
        print(f"[Error] Failed to convert HEX to JSON: {e}")
        return None

# Function to store data in Excel
def store_data_in_excel(data):
    try:
        # Extract values from the JSON data
        packet_count = data.get("packet_count", "N/A")
        timestamp = time.strftime("%Y-%m-%d %H:%M:%S")
        temperature = data.get("temperature", "N/A")
        humidity = data.get("humidity", "N/A")
        pressure = data.get("pressure", "N/A")
        device_name = data.get("device_name", "N/A")
        address = data.get("address", "N/A")
          # Get packet count from the JSON data

        # Check if the file exists
        try:
            wb = load_workbook(excel_file)
            ws = wb.active
        except FileNotFoundError:
            # Create a new workbook if the file doesn't exist
            wb = Workbook()
            ws = wb.active
            # Write header (including packet count)
            ws.append(["Packet Count","Timestamp", "Temperature", "Humidity", "Pressure", "Device Name", "Address"])
            
        # Write a new row of data (including packet count)
        ws.append([packet_count,timestamp, temperature, humidity, pressure, device_name, address])

        # Save the file
        wb.save(excel_file)

        # Print the data to the console
        print(f" Data saved to Excel: Timestamp={timestamp}, Temperature={temperature}, Humidity={humidity}, Pressure={pressure}, Device Name={device_name}, Address={address}, Packet Count={packet_count}")

    except Exception as e:
        print(f"[Error] Failed to store data in Excel: {e}")

# Main Logic
def main():
    port_name = ask_for_port()
    print(f" Opening port: {port_name}")

    try:
        ser = serial.Serial(port_name, 115200, timeout=2)
    except serial.SerialException as e:
        print(f" Failed to open port: {e}")
        sys.exit(1)

    print(" Waiting for HEX or JSON data from nRF52840...\n")

    try:
        while True:
            if ser.in_waiting > 0:
                raw_line = ser.readline().decode('utf-8', errors='ignore').strip()
                if raw_line:
                    # Check if it's HEX data
                    if all(c in '0123456789abcdefABCDEF' for c in raw_line.replace(" ", "")) and len(raw_line) % 2 == 0:
                        json_data = hex_to_json(raw_line.replace(" ", ""))
                    else:
                        # Assume it's plain JSON
                        try:
                            json_data = json.loads(raw_line)
                        except json.JSONDecodeError:
                            print(f"[Error] Invalid JSON format: {raw_line}")
                            json_data = None

                    if json_data:
                        print(f"[USB] Parsed JSON:\n{json.dumps(json_data, indent=2)}")
                        # Store in Excel
                        store_data_in_excel(json_data)
                        client.publish(publish_topic, json.dumps(json_data))
                        print(f"[MQTT] Published to {publish_topic}\n")

            time.sleep(0.2)

    except KeyboardInterrupt:
        print(" Exiting...")

    finally:
        ser.close()
        client.loop_stop()
        client.disconnect()
        print(" Serial and MQTT connection closed.")

# Entry Point
if __name__ == "__main__":
    main()
        



